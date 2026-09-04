import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename, percentage):
    wb= xl.load_workbook(filename)
    sheet= wb['Sheet1']

    for row in range(2,sheet.max_row+1):
        cell = sheet.cell(row, 3)
        corrected_price = (1-percentage/100)*cell.value
        new_cell= sheet.cell(row, 4)
        new_cell.value= corrected_price
        new_cell_title=sheet.cell(1, 4)
        new_cell_title.value = 'Corrected Price'
    values= Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart=BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'a6')
    wb.save(filename)


percent=int(input("Enter Price Reduction Percentage: "))
process_workbook("SalesList.xlsx", percent)
